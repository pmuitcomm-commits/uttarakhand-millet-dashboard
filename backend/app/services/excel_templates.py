"""Excel template generation and parsing for standardized MIS uploads."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from fastapi import HTTPException, status


@dataclass(frozen=True)
class ExcelTemplate:
    """Definition for one standardized upload workbook."""

    key: str
    filename: str
    columns: tuple[str, ...]
    required_columns: tuple[str, ...]
    instructions: tuple[str, ...]
    dropdowns: dict[str, tuple[str, ...]]


SCHEME_TYPE_OPTIONS = (
    "cultivation_input",
    "shg_intake",
    "transportation",
    "bukhari_storage",
    "sowing_incentive",
    "block_award",
)

TEMPLATES: dict[str, ExcelTemplate] = {
    "farmers_upload": ExcelTemplate(
        key="farmers_upload",
        filename="farmers_upload.xlsx",
        columns=("farmer_code", "name", "district", "block", "village", "mobile"),
        required_columns=("name", "district", "block"),
        instructions=(
            "Use one row per farmer or farmer group.",
            "farmer_code is optional. Leave blank to let the system generate or resolve a code.",
            "name, district, and block are required when creating a new farmer.",
        ),
        dropdowns={},
    ),
    "farmer_scheme_transactions_upload": ExcelTemplate(
        key="farmer_scheme_transactions_upload",
        filename="farmer_scheme_transactions_upload.xlsx",
        columns=(
            "farmer_id",
            "farmer_code",
            "name",
            "district",
            "block",
            "village",
            "mobile",
            "scheme_type",
            "millet_type",
            "quantity",
            "area_ha",
            "no_of_items",
            "production",
            "type_of_sowing",
            "incentive",
            "award",
            "transaction_date",
            "remarks",
        ),
        required_columns=("scheme_type",),
        instructions=(
            "Use farmer_id or farmer_code for existing farmers.",
            "If farmer_id and farmer_code are blank, provide name, district, and block to create or resolve the farmer.",
            "transaction_date should use YYYY-MM-DD format.",
        ),
        dropdowns={"scheme_type": SCHEME_TYPE_OPTIONS},
    ),
    "millet_production_upload": ExcelTemplate(
        key="millet_production_upload",
        filename="millet_production_upload.xlsx",
        columns=(
            "district_id",
            "block_id",
            "millet_id",
            "season_id",
            "year",
            "area_hectare",
            "production_ton",
        ),
        required_columns=("district_id", "millet_id", "year"),
        instructions=(
            "Use this workbook for aggregate analytics/statistical production data only.",
            "Do not enter farmer-level incentive rows in this workbook.",
        ),
        dropdowns={},
    ),
    "storage_processing_upload": ExcelTemplate(
        key="storage_processing_upload",
        filename="storage_processing_upload.xlsx",
        columns=(
            "district",
            "block",
            "facility_name",
            "facility_type",
            "capacity_mt",
            "operational_status",
            "remarks",
        ),
        required_columns=("facility_name",),
        instructions=(
            "Use this workbook for infrastructure, storage, and processing master data.",
            "Do not enter farmer-level incentive rows in this workbook.",
        ),
        dropdowns={
            "facility_type": ("storage", "processing", "collection_center", "other"),
            "operational_status": ("operational", "under_construction", "planned", "inactive"),
        },
    ),
}


def get_template(template_key: str) -> ExcelTemplate:
    """Return a template definition or raise a 404."""
    template = TEMPLATES.get(template_key)
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Unknown Excel template",
        )
    return template


def build_template_workbook(template: ExcelTemplate) -> bytes:
    """Build a template workbook with data, instructions, and dropdown sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel dependency openpyxl is not installed",
        ) from exc

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    instructions_sheet = workbook.create_sheet("Instructions")
    lists_sheet = workbook.create_sheet("Lists")

    header_fill = PatternFill("solid", fgColor="DDEDE7")
    required_fill = PatternFill("solid", fgColor="F8D7DA")
    for column_index, column_name in enumerate(template.columns, start=1):
        cell = data_sheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = required_fill if column_name in template.required_columns else header_fill
        data_sheet.column_dimensions[cell.column_letter].width = max(16, len(column_name) + 2)

    instructions_sheet.cell(row=1, column=1, value="Instructions").font = Font(bold=True)
    for row_index, instruction in enumerate(template.instructions, start=2):
        instructions_sheet.cell(row=row_index, column=1, value=instruction)
    instructions_sheet.cell(
        row=len(template.instructions) + 3,
        column=1,
        value=f"Required columns: {', '.join(template.required_columns) or 'None'}",
    )
    instructions_sheet.column_dimensions["A"].width = 96

    dropdown_column = 1
    for column_name, options in template.dropdowns.items():
        lists_sheet.cell(row=1, column=dropdown_column, value=column_name)
        for option_index, option in enumerate(options, start=2):
            lists_sheet.cell(row=option_index, column=dropdown_column, value=option)

        source = (
            f"=Lists!${lists_sheet.cell(row=2, column=dropdown_column).column_letter}$2:"
            f"${lists_sheet.cell(row=len(options) + 1, column=dropdown_column).column_letter}"
            f"${len(options) + 1}"
        )
        validation = DataValidation(type="list", formula1=source, allow_blank=True)
        data_sheet.add_data_validation(validation)
        if column_name in template.columns:
            column_number = template.columns.index(column_name) + 1
            column_letter = data_sheet.cell(row=1, column=column_number).column_letter
            validation.add(f"{column_letter}2:{column_letter}5000")
        dropdown_column += 1

    lists_sheet.sheet_state = "hidden"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_workbook_rows(file_bytes: bytes, filename: str, allowed_columns: tuple[str, ...]) -> list[dict[str, Any]]:
    """Parse the Data sheet from an uploaded .xlsx workbook."""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload a .xlsx file",
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel dependency openpyxl is not installed",
        ) from exc

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        ) from exc

    worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty",
        )

    allowed = set(allowed_columns)
    header = [str(value or "").strip() for value in rows[0]]
    matched_columns = [column if column in allowed else "" for column in header]
    if not any(matched_columns):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file has no recognized columns",
        )

    parsed_rows: list[dict[str, Any]] = []
    for source_row in rows[1:]:
        parsed_row: dict[str, Any] = {}
        for column_index, column_name in enumerate(matched_columns):
            if not column_name or column_index >= len(source_row):
                continue
            value = source_row[column_index]
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            parsed_row[column_name] = value
        if parsed_row:
            parsed_rows.append(parsed_row)

    if not parsed_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file does not contain non-empty data rows",
        )
    return parsed_rows

