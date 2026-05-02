"""
Runtime-reflected block data entry endpoints for scheme tables.
"""

import logging
from io import BytesIO
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field
from sqlalchemy import MetaData, Table, inspect
from sqlalchemy.exc import NoSuchTableError, SQLAlchemyError
from sqlalchemy.orm import Session

from ..database import get_db
from .auth import require_role
from .auth_roles import normalize_role

router = APIRouter(prefix="/block-data", tags=["Block Data"])
logger = logging.getLogger(__name__)

BLOCK_DATA_TABLES = {
    "inputs-millet-cultivation": "millet_cultivation_inputs",
    "incentive-sowing": "sowing_incentives",
    "bukhari-storage-structure": "bukhari_storage",
    "transportation-millet-intake": "millet_transportation_expenditure",
    "incentive-millet-intake-shg": "millet_intake_shg_incentives",
    "awards-excellent-work-block": "block_level_awards",
    "pmu-establishment-capacity": "pmu_capacity_building",
    "administrative-expenses": "administrative_expenses",
}

ALLOWED_BLOCK_DATA_TABLES = tuple(BLOCK_DATA_TABLES.values())
DISTRICT_COLUMN_CANDIDATES = ("district", "district_name")
BLOCK_COLUMN_CANDIDATES = ("block", "block_name")
MAX_EXCEL_UPLOAD_BYTES = 10 * 1024 * 1024


class BlockDataRowsRequest(BaseModel):
    """Bulk insert payload for one reflected block data table."""

    rows: list[dict[str, Any]] = Field(default_factory=list)


def _validate_table_name(table_name: str) -> str:
    """Allow dynamic table writes only for approved scheme tables."""
    if table_name not in ALLOWED_BLOCK_DATA_TABLES:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invalid block data section",
        )
    return table_name


def _get_inspector(db: Session):
    """Return the SQLAlchemy inspector bound to the active database."""
    return inspect(db.bind)


def _get_reflected_table(db: Session, table_name: str) -> Table:
    """Reflect one approved table without creating or changing schema."""
    try:
        metadata = MetaData()
        return Table(table_name, metadata, autoload_with=db.bind)
    except NoSuchTableError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Block data table not found: {table_name}",
        ) from exc


def _autoincrement_enabled(value: Any) -> bool:
    """Normalize SQLAlchemy inspector autoincrement metadata."""
    if value is True:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"auto", "true", "1", "yes"}
    return False


def _column_metadata(db: Session, table_name: str) -> list[dict[str, Any]]:
    """Read live column metadata for one approved scheme table."""
    try:
        inspector = _get_inspector(db)
        primary_key = inspector.get_pk_constraint(table_name).get("constrained_columns") or []
        primary_key_columns = set(primary_key)
        columns = inspector.get_columns(table_name)
    except NoSuchTableError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Block data table not found: {table_name}",
        ) from exc

    metadata = []
    for column in columns:
        column_name = column["name"]
        is_primary_key = column_name in primary_key_columns
        has_default = column.get("default") is not None
        has_identity = bool(column.get("identity"))
        has_computed = bool(column.get("computed"))
        is_autoincrement = _autoincrement_enabled(column.get("autoincrement"))
        is_generated = has_identity or has_computed or (
            is_primary_key and (is_autoincrement or has_default)
        )
        insertable = not is_generated

        metadata.append(
            {
                "name": column_name,
                "type": str(column.get("type")),
                "nullable": bool(column.get("nullable", True)),
                "primary_key": is_primary_key,
                "has_default": has_default or has_identity or has_computed,
                "insertable": insertable,
                "required": insertable
                and not bool(column.get("nullable", True))
                and not (has_default or has_identity or has_computed),
            }
        )

    return metadata


def _find_column(columns: list[str], candidates: tuple[str, ...]) -> Optional[str]:
    """Find a column name by case-insensitive candidate list."""
    by_lower = {column.lower(): column for column in columns}
    for candidate in candidates:
        found = by_lower.get(candidate.lower())
        if found:
            return found
    return None


def _clean_filter_value(value: Optional[str]) -> Optional[str]:
    """Normalize optional district/block query filters."""
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _normalize_payload_key(value: Any) -> str:
    """Normalize client or Excel column names for loose matching."""
    return "".join(
        character.lower()
        for character in str(value or "").strip()
        if character.isalnum()
    )


def _is_blank(value: Any) -> bool:
    """Treat None and whitespace-only strings as empty cell values."""
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_row(raw_row: dict[str, Any], insertable_columns: list[str]) -> dict[str, Any]:
    """Keep only live table columns from a submitted row."""
    columns_by_key = {
        _normalize_payload_key(column_name): column_name
        for column_name in insertable_columns
    }
    normalized_row: dict[str, Any] = {}

    for raw_key, value in raw_row.items():
        matched_column = columns_by_key.get(_normalize_payload_key(raw_key))
        if matched_column and not _is_blank(value):
            normalized_row[matched_column] = value

    return normalized_row


def _apply_scope_defaults(
    row: dict[str, Any],
    insertable_columns: list[str],
    current_user,
    district: Optional[str],
    block: Optional[str],
) -> dict[str, Any]:
    """Fill existing district/block columns from role scope or admin filters."""
    scoped_row = dict(row)
    district_column = _find_column(insertable_columns, DISTRICT_COLUMN_CANDIDATES)
    block_column = _find_column(insertable_columns, BLOCK_COLUMN_CANDIDATES)
    selected_district = _clean_filter_value(district)
    selected_block = _clean_filter_value(block)
    current_role = normalize_role(current_user.get("role"))

    if district_column:
        district_value = (
            current_user["district"]
            if current_role == "block"
            else selected_district or scoped_row.get(district_column)
        )
        if not _is_blank(district_value):
            scoped_row[district_column] = district_value

    if block_column:
        block_value = (
            current_user["block"]
            if current_role == "block"
            else selected_block or scoped_row.get(block_column)
        )
        if not _is_blank(block_value):
            scoped_row[block_column] = block_value

    return scoped_row


def _cell_to_payload_value(value: Any) -> Any:
    """Normalize spreadsheet cell values for JSON responses and inserts."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _rows_from_xlsx(file_bytes: bytes) -> list[list[Any]]:
    """Read the first worksheet from an .xlsx file."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel parser dependency openpyxl is not installed",
        ) from exc

    workbook = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _rows_from_xls(file_bytes: bytes) -> list[list[Any]]:
    """Read the first worksheet from an .xls file."""
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel parser dependency xlrd is not installed",
        ) from exc

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    worksheet = workbook.sheet_by_index(0)
    return [
        [worksheet.cell_value(row_index, column_index) for column_index in range(worksheet.ncols)]
        for row_index in range(worksheet.nrows)
    ]


def _parse_excel_rows(
    file_bytes: bytes,
    filename: str,
    insertable_columns: list[str],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    """Parse one Excel worksheet and keep only columns that exist in the table."""
    lower_filename = filename.lower()
    try:
        if lower_filename.endswith(".xlsx"):
            worksheet_rows = _rows_from_xlsx(file_bytes)
        elif lower_filename.endswith(".xls"):
            worksheet_rows = _rows_from_xls(file_bytes)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Upload a .xlsx or .xls file",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        ) from exc

    if not worksheet_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty",
        )

    header_row = worksheet_rows[0]
    columns_by_key = {
        _normalize_payload_key(column_name): column_name
        for column_name in insertable_columns
    }
    matched_headers: list[Optional[str]] = []
    matched_columns: list[str] = []
    ignored_columns: list[str] = []

    for header in header_row:
        header_label = str(header or "").strip()
        matched_column = columns_by_key.get(_normalize_payload_key(header_label))
        matched_headers.append(matched_column)
        if matched_column:
            matched_columns.append(matched_column)
        elif header_label:
            ignored_columns.append(header_label)

    if not matched_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file has no columns matching this database table",
        )

    parsed_rows: list[dict[str, Any]] = []
    for worksheet_row in worksheet_rows[1:]:
        parsed_row: dict[str, Any] = {}
        for column_index, column_name in enumerate(matched_headers):
            if not column_name or column_index >= len(worksheet_row):
                continue
            value = _cell_to_payload_value(worksheet_row[column_index])
            if not _is_blank(value):
                parsed_row[column_name] = value
        if parsed_row:
            parsed_rows.append(parsed_row)

    if not parsed_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file does not contain non-empty data rows",
        )

    return parsed_rows, sorted(set(matched_columns)), ignored_columns


def _prepare_insert_rows(
    table_name: str,
    column_info: list[dict[str, Any]],
    request_rows: list[dict[str, Any]],
    current_user,
    district: Optional[str],
    block: Optional[str],
) -> list[dict[str, Any]]:
    """Validate and prepare rows for insertion without mutating the database."""
    if not request_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one row is required",
        )

    insertable_columns = [
        column["name"]
        for column in column_info
        if column["insertable"]
    ]
    if not insertable_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{table_name} does not have insertable columns",
        )

    required_columns = [
        column["name"]
        for column in column_info
        if column["required"]
    ]
    district_column = _find_column(insertable_columns, DISTRICT_COLUMN_CANDIDATES)
    block_column = _find_column(insertable_columns, BLOCK_COLUMN_CANDIDATES)
    selected_district = _clean_filter_value(district)
    selected_block = _clean_filter_value(block)
    current_role = normalize_role(current_user.get("role"))
    prepared_rows: list[dict[str, Any]] = []
    validation_errors: list[str] = []

    for row_index, raw_row in enumerate(request_rows, start=1):
        if not isinstance(raw_row, dict):
            validation_errors.append(f"Row {row_index} must be an object")
            continue

        normalized_row = _normalize_row(raw_row, insertable_columns)
        user_data_columns = [
            column
            for column in normalized_row
            if column not in {district_column, block_column}
        ]
        if not user_data_columns:
            continue

        if district_column:
            district_value = (
                current_user["district"]
                if current_role == "block"
                else selected_district or normalized_row.get(district_column)
            )
            if _is_blank(district_value):
                validation_errors.append(f"District is required on row {row_index}")
            else:
                normalized_row[district_column] = district_value

        if block_column:
            block_value = (
                current_user["block"]
                if current_role == "block"
                else selected_block or normalized_row.get(block_column)
            )
            if _is_blank(block_value):
                validation_errors.append(f"Block is required on row {row_index}")
            else:
                normalized_row[block_column] = block_value

        missing_required = [
            column
            for column in required_columns
            if _is_blank(normalized_row.get(column))
        ]
        if missing_required:
            validation_errors.append(
                f"Missing required column(s) on row {row_index}: {', '.join(missing_required)}"
            )
            continue

        prepared_rows.append(normalized_row)

    if validation_errors:
        logger.warning(
            "Rejected block data insert for %s: %s",
            table_name,
            "; ".join(validation_errors),
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="; ".join(validation_errors[:5]),
        )

    if not prepared_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No non-empty rows to save",
        )

    return prepared_rows


@router.get("/{table_name}/schema")
def get_block_data_table_schema(
    table_name: str,
    current_user=Depends(require_role("admin", "block")),
    db: Session = Depends(get_db),
):
    """
    Return live column metadata for an approved block data table.
    """
    validated_table_name = _validate_table_name(table_name)
    columns = _column_metadata(db, validated_table_name)
    column_names = [column["name"] for column in columns]

    return {
        "table_name": validated_table_name,
        "columns": columns,
        "district_column": _find_column(column_names, DISTRICT_COLUMN_CANDIDATES),
        "block_column": _find_column(column_names, BLOCK_COLUMN_CANDIDATES),
    }


@router.post("/{table_name}/upload")
async def parse_block_data_excel(
    table_name: str,
    file: UploadFile = File(...),
    district: Optional[str] = Query(default=None),
    block: Optional[str] = Query(default=None),
    current_user=Depends(require_role("admin", "block")),
    db: Session = Depends(get_db),
):
    """
    Parse an Excel file and return rows filtered to live table columns.
    """
    validated_table_name = _validate_table_name(table_name)
    columns = _column_metadata(db, validated_table_name)
    insertable_columns = [
        column["name"]
        for column in columns
        if column["insertable"]
    ]
    if not insertable_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{validated_table_name} does not have insertable columns",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty",
        )
    if len(file_bytes) > MAX_EXCEL_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is too large",
        )

    parsed_rows, matched_columns, ignored_columns = _parse_excel_rows(
        file_bytes,
        file.filename or "",
        insertable_columns,
    )
    scoped_rows = [
        _apply_scope_defaults(
            row,
            insertable_columns,
            current_user,
            district,
            block,
        )
        for row in parsed_rows
    ]

    return jsonable_encoder(
        {
            "table_name": validated_table_name,
            "rows": scoped_rows,
            "matched_columns": matched_columns,
            "ignored_columns": ignored_columns,
        }
    )


@router.post("/{table_name}")
def save_block_data_rows(
    table_name: str,
    request: BlockDataRowsRequest,
    district: Optional[str] = Query(default=None),
    block: Optional[str] = Query(default=None),
    current_user=Depends(require_role("admin", "block")),
    db: Session = Depends(get_db),
):
    """
    Insert rows into one approved block data table using reflected columns only.
    """
    validated_table_name = _validate_table_name(table_name)
    table = _get_reflected_table(db, validated_table_name)
    columns = _column_metadata(db, validated_table_name)
    prepared_rows = _prepare_insert_rows(
        validated_table_name,
        columns,
        request.rows,
        current_user,
        district,
        block,
    )

    inserted_rows = []
    try:
        for row in prepared_rows:
            result = db.execute(
                table.insert()
                .values(row)
                .returning(*table.columns)
            )
            inserted = result.mappings().first()
            inserted_rows.append(dict(inserted) if inserted else row)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        logger.exception("Unable to save block data rows for %s", validated_table_name)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to save block data",
        ) from exc

    return jsonable_encoder(
        {
            "table_name": validated_table_name,
            "inserted_count": len(inserted_rows),
            "rows": inserted_rows,
        }
    )
